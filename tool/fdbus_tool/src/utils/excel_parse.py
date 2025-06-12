#!/usr/bin/env python3

"""Import openpyxl for excel file parsing."""
import sys
import logging
import openpyxl


class ExeclParser:
    """Class parse excel file and generate json data."""

    def __init__(self, xlsx):
        self.workbook = openpyxl.load_workbook(filename=xlsx)

    def __get_col_idx(self, sheet, col_name):
        for col in sheet.iter_cols():
            if col[0].value == col_name:
                return col[0].col_idx - 1
        return None

    def __get_col_idx_by_row(self, sheet, start_row, col_name):
        for col in sheet.iter_cols():
            if col[start_row].value == col_name:
                return col[start_row].col_idx - 1
        return None

    def __get_row_idx_by_col(self, sheet, target_col, row_name):
        for row in sheet.iter_rows():
            if row[target_col].value == row_name:
                return row[0].row
        return None

    def __get_sub_col_idx(self, sheet, col_name, target_row, sub_col_name):
        start_col = -1
        for col in sheet.iter_cols():
            if col[0].value == col_name:
                start_col = col[0].col_idx - 1
                break
        if start_col == -1:
            print("not find col_name:", col_name)
            return None
        for col in sheet.iter_cols(min_col=start_col):
            if col[target_row].value == sub_col_name:
                return col[target_row].col_idx - 1
        return None

    def __get_sub_col_idx_by_row(
        self, sheet, col_name, start_row, target_row, sub_col_name
    ):
        start_col = -1
        for col in sheet.iter_cols():
            if col[start_row].value == col_name:
                start_col = col[start_row].col_idx - 1
                break
        if start_col == -1:
            print("not find col_name:", col_name)
            return None
        for col in sheet.iter_cols(min_col=start_col):
            if col[target_row].value == sub_col_name:
                return col[target_row].col_idx - 1
        return None

    def __get_three_sub_col_idx(
        self, sheet, col_name, start_row, first_row, first_col_name, second_row, second_col_name
    ):
        start_col = -1
        for col in sheet.iter_cols():
            if col[start_row].value == col_name:
                start_col = col[start_row].col_idx - 1
                break
        if start_col == -1:
            print("not find col_name:", col_name)
            return None

        first_col = -1
        for col in sheet.iter_cols(min_col=start_col+1):
            if col[first_row].value == first_col_name:
                first_col = col[first_row].col_idx - 1
                break
        if first_col == -1:
            print("not find first_col_name:", first_col_name)
            return None

        for col in sheet.iter_cols(min_col=first_col+1):
            if col[second_row].value == second_col_name:
                
                return col[second_row].col_idx - 1
        return None

    def gen_data_type_json(self, sheet_name, json):
        """Generate data type json data."""
        sheet_info = self.workbook[sheet_name]
        start_row = 2

        service_name_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Service Name*")
        if service_name_idx is None:
            logging.error("Service Name* not find in Sheet")
            sys.exit(1)

        namespace_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Namespace*")
        if namespace_idx is None:
            logging.error("Namespace* not find in Sheet")
            sys.exit(1)

        data_type_name_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Data Type Name*")
        if data_type_name_idx is None:
            logging.error("Data Type Name* not find in Sheet")
            sys.exit(1)

        category_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Data Type Category*")
        if category_idx is None:
            logging.error("Data Type Category* not find in Sheet")
            sys.exit(1)

        member_name_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Member Name*")
        if member_name_idx is None:
            logging.error("Member Name* not find in Sheet")
            sys.exit(1)

        member_type_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Member Data Type Reference*")
        if member_type_idx is None:
            logging.error("Member Data Type Reference* not find in Sheet")
            sys.exit(1)

        cur_name = ""
        # 遍历行
        msg_idx = 0
        for row in sheet_info.iter_rows(min_row=start_row+2, values_only=True):
            if row[service_name_idx] is not None:
                json[row[service_name_idx]] = {
                    "ns": "",
                    "messages": [],
                }
                cur_name = row[service_name_idx]
                msg_idx = 0
            if row[namespace_idx] is not None:
                json[cur_name]["ns"] = row[namespace_idx]
            if row[data_type_name_idx] is not None:
                json[cur_name]["messages"].append(
                    {"name": row[data_type_name_idx], "category": "", "members":[]}
                )
                msg_idx = msg_idx + 1
            if row[category_idx] is not None:
                json[cur_name]["messages"][msg_idx - 1]["category"] = row[category_idx]
            if row[member_name_idx] is not None and row[member_type_idx] is not None:
                json[cur_name]["messages"][msg_idx - 1]["members"].append(
                    {"name": row[member_name_idx], "type": row[member_type_idx]}
                )
        return json

    def gen_srv_if_json(self, sheet_name, json):
        """Generate service interface json data."""
        sheet_info = self.workbook[sheet_name]

        start_row = 2
        srv_name_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Service Name*")
        if srv_name_idx is None:
            logging.error("Service Name* not find in Sheet")
            sys.exit(1)

        ns_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Namespace*")
        if ns_idx is None:
            logging.error("Namespace* not find in Sheet")
            sys.exit(1)

        if_name_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Element Name*")
        if if_name_idx is None:
            logging.error("Element Name* not find in Sheet")
            sys.exit(1)

        if_type_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Element Type*")
        if if_type_idx is None:
            logging.error("Element Type* not find in Sheet")
            sys.exit(1)

        msg_id_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Msg Id*")
        if msg_id_idx is None:
            logging.error("Msg Id* not find in Sheet")
            sys.exit(1)

        target_row = start_row + 2
        param_name_idx = self.__get_sub_col_idx_by_row(
            sheet_info, "Param", start_row, target_row, "Name*"
        )
        if param_name_idx is None:
            logging.error("Param Name* not find in Sheet")
            sys.exit(1)

        param_type_idx = self.__get_sub_col_idx_by_row(
            sheet_info, "Param", start_row, target_row, "Type*"
        )
        if param_type_idx is None:
            logging.error("Param Type* not find in Sheet")
            sys.exit(1)

        param_direction_idx = self.__get_sub_col_idx_by_row(
            sheet_info, "Param", start_row, target_row, "Direction*"
        )
        if param_direction_idx is None:
            logging.error("Param Direction* not find in Sheet")
            sys.exit(1)

        cur_srv_name = ""
        if_idx = 0
        cur_if_idx = 0
        last_if_name = ""
        data_row = target_row + 1
        # 遍历行
        for row in sheet_info.iter_rows(min_row=data_row, values_only=True):
            if row[srv_name_idx] is not None:
                json[row[srv_name_idx]] = {
                    "ns": row[ns_idx],
                    "interfaces": [],
                }
                last_if_name = ""
                if_idx = 0
                cur_srv_name = row[srv_name_idx]
            if row[if_name_idx] is not None and row[if_type_idx] is not None:
                if row[if_type_idx] == "event":
                    if_idx = if_idx + 1
                    json[cur_srv_name]["interfaces"].append(
                        {
                            "name": row[if_name_idx],
                            "type": "event",
                            "msg_id": row[msg_id_idx],
                            "param": {
                                "out": {"name": "data", "type": row[param_type_idx]}
                            }
                        }
                    )
                elif (
                    row[if_type_idx] == "rr & method"
                    or row[if_type_idx] == "ff & method"
                ):
                    if (
                        row[if_name_idx] is not None
                        and last_if_name != row[if_name_idx]
                    ):
                        last_if_name = row[if_name_idx]
                        cur_if_idx = if_idx
                        if_idx = if_idx + 1
                        method_type = "ff"
                        if row[if_type_idx] == "rr & method":
                            method_type = "rr"

                        json[cur_srv_name]["interfaces"].append(
                            {
                                "name": row[if_name_idx],
                                "type": "method",
                                "msg_id": row[msg_id_idx],
                                "mtd_type": method_type,
                                "param": {
                                    "has_param": True,
                                    "in": {"name": "", "type": ""},
                                    "out": {"name": "", "type": ""},
                                },
                            }
                        )
                    if row[param_direction_idx] == "IN":
                        json[cur_srv_name]["interfaces"][cur_if_idx]["param"]["in"] = {
                            "name": row[param_name_idx],
                            "type": row[param_type_idx],
                        }
                    elif row[param_direction_idx] == "OUT":
                        json[cur_srv_name]["interfaces"][cur_if_idx]["param"]["out"] = {
                            "name": row[param_name_idx],
                            "type": row[param_type_idx],
                        }
                    elif row[param_direction_idx] == "NONE":
                        json[cur_srv_name]["interfaces"][cur_if_idx]["param"]["has_param"] = False
        return json

    def gen_deploy_json(self, sheet_name, json):
        """Generate service deployment json data."""
        sheet_info = self.workbook[sheet_name]

        start_row = 2
        deploy_ecu_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Deploy ECU*")
        if deploy_ecu_idx is None:
            logging.error("Deploy ECU* not find in Sheet")
            sys.exit(1)

        ip_idx = self.__get_col_idx_by_row(sheet_info, start_row, "IP Address*")
        if ip_idx is None:
            logging.error("IP Address* not find in Sheet")
            sys.exit(1)

        app_name_idx = self.__get_col_idx_by_row(sheet_info, start_row, "App Name*")
        if app_name_idx is None:
            logging.error("App Name* not find in Sheet")
            sys.exit(1)

        app_role_idx = self.__get_col_idx_by_row(
            sheet_info, start_row, "Server/Client*"
        )
        if app_role_idx is None:
            logging.error("Server/Client* not find in Sheet")
            sys.exit(1)

        srv_name_idx = self.__get_col_idx_by_row(sheet_info, start_row, "Service Name*")
        if srv_name_idx is None:
            logging.error("Service Name* not find in Sheet")
            sys.exit(1)

        cur_ecu_name = ""
        idx = 0
        data_row = start_row + 3
        # 遍历行
        for row in sheet_info.iter_rows(min_row=data_row, values_only=True):
            if row[deploy_ecu_idx] is not None and row[ip_idx] is not None:
                json[row[deploy_ecu_idx]] = {
                    "ip": row[ip_idx],
                    "apps": [],
                }
                cur_ecu_name = row[deploy_ecu_idx]
                idx = 0
            if row[app_name_idx] is not None :
                json[cur_ecu_name]["apps"].append(
                    {
                        "name": row[app_name_idx],
                        "services": [],
                        "req_services": [],
                    }
                )
                idx = idx + 1
            if (
                row[app_role_idx] is not None
                and row[srv_name_idx] is not None
            ):
                if row[app_role_idx] == "server":
                    json[cur_ecu_name]["apps"][idx - 1]["services"].append(
                        {
                            "srv_name": row[srv_name_idx]
                        }
                    )
                if row[app_role_idx] == "client":
                    json[cur_ecu_name]["apps"][idx - 1]["req_services"].append(
                        {
                            "srv_name": row[srv_name_idx]
                        }
                    )
        return json

    def close_excel(self):
        """close excel file."""
        self.workbook.close()
